import easyocr
import re
import pandas as pd
from openpyxl import load_workbook
import cv2     

# =====================================================
# PATHS
# =====================================================
IMAGE_PATH = r"C:\Drive_d\Python\F-AI\T4\Input\images\invoice1_page_1.png"
EXCEL_PATH = r"C:\Drive_d\Python\F-AI\T4\Output Template.xlsx"

TABLE_SHEET  = "Table_1"
HEADER_SHEET = "Invoice_Header"

# =====================================================
# REGEX
# =====================================================
NUM_RE   = r"\d+\.\d+"
TOTAL_RE = r"TOTAL[:\s]*\{?(\d+\.\d+)\s*/\s*(\d+\.\d+)"

# =====================================================
# OCR INIT
# =====================================================
reader = easyocr.Reader(["en"], gpu=False)
lines = [l.strip() for l in reader.readtext(IMAGE_PATH, detail=0) if l.strip()]

# =====================================================
# HELPER
# =====================================================
def clean_price(val: str) -> float:
    # removes leading OCR noise like 8218.75 → 218.75
    if val.startswith("8") and len(val) > 6:
        val = val[1:]
    return float(val)

def grab(pattern, text):
    m = re.search(pattern, text, re.I | re.S)
    return m.group(1).strip() if m else ""

# =====================================================
# ================= TABLE_1 (UNCHANGED) ===============
# =====================================================
total_tax = 0.0
total_amount = 0.0

for i, l in enumerate(lines):
    if "TOTAL" in l:
        combo = l + " " + (lines[i + 1] if i + 1 < len(lines) else "")
        m = re.search(TOTAL_RE, combo)
        if m:
            total_tax = clean_price(m.group(1))
            total_amount = clean_price(m.group(2)[1:])
        break

start_idx = None
for i, l in enumerate(lines):
    if "Amount-Amount" in l.replace(" ", ""):
        start_idx = i + 1
        break

if start_idx is None:
    raise RuntimeError("❌ 'Amount-Amount' anchor not found")

description_lines = []
numbers = []

for l in lines[start_idx:]:
    if re.search(r"(TOTAL|Amount in Words|Authorized|HSN|Shipping)", l, re.I):
        break

    nums = re.findall(NUM_RE, l)
    if nums:
        numbers.extend(nums)
    else:
        description_lines.append(l)

if len(numbers) < 3:
    raise RuntimeError("❌ OCR numbers incomplete")

unit_price = clean_price(numbers[0])
tax_amount = clean_price(numbers[-2])
total_amt  = clean_price(numbers[-1])
net_amount = round(total_amt - tax_amount, 2)

rows = [[
    1,
    " ".join(description_lines).strip(),
    unit_price,
    0.0,
    1,
    net_amount,
    "12%",
    "IGST",
    tax_amount,
    total_amt
]]

df = pd.DataFrame(rows, columns=[
    "Sl.No",
    "Description",
    "UnitPrice",
    "Discount",
    "Qty",
    "NetAmount",
    "TaxRate",
    "TaxType",
    "TaxAmount",
    "TotalAmount"
])

print(df)

# =====================================================
# WRITE TABLE_1
# =====================================================
wb = load_workbook(EXCEL_PATH)
ws_table = wb[TABLE_SHEET]

start_row = 2
for r, row in enumerate(df.itertuples(index=False), start_row):
    for c, val in enumerate(row, 1):
        ws_table.cell(row=r, column=c).value = val

total_row = start_row + len(df)
ws_table.cell(row=total_row, column=2).value = "TOTAL:"
ws_table.cell(row=total_row, column=9).value = total_tax
ws_table.cell(row=total_row, column=10).value = total_amount

# =====================================================
# ================= INVOICE_HEADER ====================
# =====================================================
img = cv2.imread(IMAGE_PATH)
h, w, _ = img.shape
left_img  = img[:, :w//2]
right_img = img[:, w//2:]

left_text  = "\n".join([t for _, t, _ in reader.readtext(left_img)])
right_text = "\n".join([t for _, t, _ in reader.readtext(right_img)])

full_text = left_text + "\n" + right_text

header_data = {
    "billing_address": grab(r"Billing Address\s*(.*?)\s*Shipping Address", full_text),
    "shipping_address": grab(r"Shipping Address\s*(.*?)\s*Invoice Number", full_text),
    "invoice_type": grab(r"(Tax Invoice/Bill of Supply/Cash Memo)", full_text),
    "order_number": grab(r"Order Number[:\s]*([\w\-]+)", full_text),
    "invoice_number": grab(r"Invoice Number[:\s]*([\w\-]+)", full_text),
    "order_date": grab(r"Order Date[:\s]*([\d\.]+)", full_text),
    "invoice_details": grab(r"Invoice Details\s*(.*?)\s*Invoice Date", full_text),
    "invoice_date": grab(r"Invoice Date\s*([\d\.]+)", full_text),
    "seller_info": grab(r"Sold By\s*(.*?)\s*PAN No", full_text),
    "seller_pan": grab(r"PAN No[:\s]*([A-Z0-9]+)", full_text),
    "seller_gst": grab(r"GST Registration No[:\s]*([A-Z0-9]+)", full_text),
    "fssai_license": grab(r"FSSAI[:\s]*([\d]+)", full_text),
    "billing_state_code": grab(r"State/UT Code[:\s]*(\d+)", full_text),
    "shipping_state_code": grab(r"StatelUT Code[:\s]*(\d+)", full_text),
    "place_of_supply": grab(r"Place of supply[:\s]*([A-Z\s]+)", full_text),
    "place_of_delivery": grab(r"Place of delivery[:\s]*([A-Z\s]+)", full_text),
    "reverse_charge": grab(r"reverse charge\s*(Yes|No)", full_text),
    "amount_in_words": grab(r"Amount in Words[:\s]*(.*?)\s*For", full_text),
    "seller_name": grab(r"Sold By\s*([A-Z\s]+)", full_text),
    "seller_address": grab(r"Sold By\s*.*?\n(.*?)\nIN", full_text),
    "total_tax": total_tax,
    "total_amount": total_amount
}

# =====================================================
# WRITE INVOICE_HEADERimport easyocr
import re
import pandas as pd
from openpyxl import load_workbook
import cv2
import os     

# =====================================================
# PATHS
# =====================================================
IMAGE_PATH = r"C:\Drive_d\Python\F-AI\T4\Input\images\invoice1_page_1.png"
TEMPLATE_EXCEL_PATH = r"C:\Drive_d\Python\F-AI\T4\Output Template.xlsx"
OUTPUT_DIR = r"C:\Drive_d\Python\F-AI\T4\Outputs\Amazon"
OUTPUT_EXCEL_PATH = os.path.join(OUTPUT_DIR, "Amazon_Invoice_Output.xlsx")

TABLE_SHEET  = "Table_1"
HEADER_SHEET = "Invoice_Header"

os.makedirs(OUTPUT_DIR, exist_ok=True)

# =====================================================
# REGEX
# =====================================================
NUM_RE   = r"\d+\.\d+"
TOTAL_RE = r"TOTAL[:\s]*\{?(\d+\.\d+)\s*/\s*(\d+\.\d+)"

# =====================================================
# OCR INIT
# =====================================================
reader = easyocr.Reader(["en"], gpu=False)
lines = [l.strip() for l in reader.readtext(IMAGE_PATH, detail=0) if l.strip()]

# =====================================================
# HELPERS
# =====================================================
def clean_price(val: str) -> float:
    if val.startswith("8") and len(val) > 6:
        val = val[1:]
    return float(val)

def grab(pattern, text):
    m = re.search(pattern, text, re.I | re.S)
    return m.group(1).strip() if m else ""

# =====================================================
# ================= TABLE_1 ===========================
# =====================================================
total_tax = 0.0
total_amount = 0.0

for i, l in enumerate(lines):
    if "TOTAL" in l:
        combo = l + " " + (lines[i + 1] if i + 1 < len(lines) else "")
        m = re.search(TOTAL_RE, combo)
        if m:
            total_tax = clean_price(m.group(1))
            total_amount = clean_price(m.group(2)[1:])
        break

start_idx = None
for i, l in enumerate(lines):
    if "Amount-Amount" in l.replace(" ", ""):
        start_idx = i + 1
        break

if start_idx is None:
    raise RuntimeError("❌ 'Amount-Amount' anchor not found")

description_lines = []
numbers = []

for l in lines[start_idx:]:
    if re.search(r"(TOTAL|Amount in Words|Authorized|HSN|Shipping)", l, re.I):
        break

    nums = re.findall(NUM_RE, l)
    if nums:
        numbers.extend(nums)
    else:
        description_lines.append(l)

if len(numbers) < 3:
    raise RuntimeError("❌ OCR numbers incomplete")

unit_price = clean_price(numbers[0])
tax_amount = clean_price(numbers[-2])
total_amt  = clean_price(numbers[-1])
net_amount = round(total_amt - tax_amount, 2)

rows = [[
    1,
    " ".join(description_lines).strip(),
    unit_price,
    0.0,
    1,
    net_amount,
    "12%",
    "IGST",
    tax_amount,
    total_amt
]]

df = pd.DataFrame(rows, columns=[
    "Sl.No",
    "Description",
    "UnitPrice",
    "Discount",
    "Qty",
    "NetAmount",
    "TaxRate",
    "TaxType",
    "TaxAmount",
    "TotalAmount"
])

print(df)

# =====================================================
# WRITE TABLE_1
# =====================================================
wb = load_workbook(TEMPLATE_EXCEL_PATH)
ws_table = wb[TABLE_SHEET]

start_row = 2
for r, row in enumerate(df.itertuples(index=False), start_row):
    for c, val in enumerate(row, 1):
        ws_table.cell(row=r, column=c).value = val

total_row = start_row + len(df)
ws_table.cell(row=total_row, column=2).value = "TOTAL:"
ws_table.cell(row=total_row, column=9).value = total_tax
ws_table.cell(row=total_row, column=10).value = total_amount

# =====================================================
# ================= INVOICE_HEADER ====================
# =====================================================
img = cv2.imread(IMAGE_PATH)
h, w, _ = img.shape
left_img  = img[:, :w//2]
right_img = img[:, w//2:]

left_text  = "\n".join([t for _, t, _ in reader.readtext(left_img)])
right_text = "\n".join([t for _, t, _ in reader.readtext(right_img)])
full_text = left_text + "\n" + right_text

header_data = {
    "billing_address": grab(r"Billing Address\s*(.*?)\s*Shipping Address", full_text),
    "shipping_address": grab(r"Shipping Address\s*(.*?)\s*Invoice Number", full_text),
    "invoice_type": grab(r"(Tax Invoice/Bill of Supply/Cash Memo)", full_text),
    "order_number": grab(r"Order Number[:\s]*([\w\-]+)", full_text),
    "invoice_number": grab(r"Invoice Number[:\s]*([\w\-]+)", full_text),
    "order_date": grab(r"Order Date[:\s]*([\d\.]+)", full_text),
    "invoice_details": grab(r"Invoice Details\s*(.*?)\s*Invoice Date", full_text),
    "invoice_date": grab(r"Invoice Date\s*([\d\.]+)", full_text),
    "seller_info": grab(r"Sold By\s*(.*?)\s*PAN No", full_text),
    "seller_pan": grab(r"PAN No[:\s]*([A-Z0-9]+)", full_text),
    "seller_gst": grab(r"GST Registration No[:\s]*([A-Z0-9]+)", full_text),
    "fssai_license": grab(r"FSSAI[:\s]*([\d]+)", full_text),
    "billing_state_code": grab(r"State/UT Code[:\s]*(\d+)", full_text),
    "shipping_state_code": grab(r"StatelUT Code[:\s]*(\d+)", full_text),
    "place_of_supply": grab(r"Place of supply[:\s]*([A-Z\s]+)", full_text),
    "place_of_delivery": grab(r"Place of delivery[:\s]*([A-Z\s]+)", full_text),
    "reverse_charge": grab(r"reverse charge\s*(Yes|No)", full_text),
    "amount_in_words": grab(r"Amount in Words[:\s]*(.*?)\s*For", full_text),
    "seller_name": grab(r"Sold By\s*([A-Z\s]+)", full_text),
    "seller_address": grab(r"Sold By\s*.*?\n(.*?)\nIN", full_text),
    "total_tax": total_tax,
    "total_amount": total_amount
}

ws_header = wb[HEADER_SHEET]
row = 2
for k, v in header_data.items():
    ws_header.cell(row=row, column=1).value = k
    ws_header.cell(row=row, column=2).value = v
    row += 1

wb.save(OUTPUT_EXCEL_PATH)

print("✅ AMAZON INVOICE — TABLE + HEADER COMPLETED")
print("📄 Saved to:", OUTPUT_EXCEL_PATH)

# =====================================================
ws_header = wb[HEADER_SHEET]
row = 2
for k, v in header_data.items():
    ws_header.cell(row=row, column=1).value = k
    ws_header.cell(row=row, column=2).value = v
    row += 1

wb.save(EXCEL_PATH)

print("✅ AMAZON INVOICE — TABLE + HEADER COMPLETED")


# img = cv2.imread(IMAGE_PATH)
# h, w, _ = img.shape
# left_col  = img[:, :w//2]
# right_col = img[:, w//2:]
# print("left col")
# Left_results = reader.readtext(left_col)
# for bbox, text, conf in Left_results:
#     print(f"{text}")

# print("\nright col")
# Right_results = reader.readtext(right_col)      
# for bbox, text, conf in Right_results:
#     print(f"{text}")