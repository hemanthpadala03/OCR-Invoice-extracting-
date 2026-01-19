import easyocr
import re
import pandas as pd
from openpyxl import load_workbook

IMAGE_PATH = r"C:\Drive_d\Python\F-AI\T4\Input\images\Invoice_7127787656_page_1.png"
TEMPLATE_PATH = r"C:\Drive_d\Python\F-AI\T4\Outputs\Zomato\Output Template.xlsx"

reader = easyocr.Reader(['en'], gpu=False)
results = reader.readtext(IMAGE_PATH)

lines = [t.strip() for _, t, _ in results if t.strip()]
full_text = "\n".join(lines)

def grab(pattern, text):
    m = re.search(pattern, text, re.I | re.S)
    return m.group(1).strip() if m else ""

def is_number(x):
    try:
        float(x)
        return True
    except:
        return False

def clean(x):
    return re.sub(r"\s+", " ", x).strip()

# ================= HEADER EXTRACTION =================
data = {}

data["invoice_type"] = "Tax Invoice"
data["invoice_number"] = grab(r"Invoice\s*No\.?\s*[:\-]?\s*([A-Z0-9]+)", full_text)
data["invoice_date"] = grab(r"Invoice\s*Date\s*[:\-]?\s*([\d\/\-]+)", full_text)
data["order_date"] = data["invoice_date"]
data["order_number"] = grab(r"Order\s*ID\s*[:\-]?\s*(\d+)", full_text)

data["billing_address"] = clean(
    grab(r"Delivery\s*Address\s*(.*?)(State\s*name|HSN\s*Code)", full_text)
)
data["shipping_address"] = data["billing_address"]

data["place_of_supply"] = grab(
    r"State\s*name\s*&\s*Place\s*of\s*Supply\s*:\s*([A-Za-z ]+)", full_text
)
data["place_of_delivery"] = data["place_of_supply"]

state_code = grab(r"\((\d{2})\)", full_text)
data["billing_state_code"] = state_code
data["shipping_state_code"] = state_code

data["seller_name"] = grab(r"Legal\s*Entity\s*Name\s*(.*)", full_text)
data["seller_address"] = clean(
    grab(r"Restaurant\s*Address\s*(.*?)(Restaurant\s*GSTIN)", full_text)
)

data["seller_gst"] = grab(r"Restaurant\s*GSTIN\s*([0-9A-Z]{15})", full_text)
data["seller_pan"] = data["seller_gst"][2:12] if data["seller_gst"] else ""
data["fssai_license"] = grab(r"Restaurant\s*FSSAI\s*(\d+)", full_text)

data["seller_info"] = f"{data['seller_name']} | GST: {data['seller_gst']}"
data["reverse_charge"] = grab(r"Reverse\s*charge\s*(Yes|No)", full_text)
data["invoice_details"] = "Zomato Tax Invoice"

data["amount_in_words"] = clean(
    grab(r"Amount\s*\(in\s*words\)\s*:\s*(.*?)Only", full_text)
)

data["total_amount"] = grab(r"Total\s*Value\s*([\d\.]+)", full_text)

try:
    data["total_tax"] = round(float(data["total_amount"]) * 0.05, 2)
except:
    data["total_tax"] = ""

# ================= LINE ITEM EXTRACTION =================
items = []
i = 0
n = len(lines)

while i < n and lines[i].lower() != "particulars":
    i += 1
i += 1

SKIP_WORDS = {"gross", "discount", "net", "rate", "inr", "total", "value"}

while i < n:
    if lines[i].lower().startswith("total value"):
        break

    if any(w in lines[i].lower() for w in SKIP_WORDS):
        i += 1
        continue

    description = lines[i]
    i += 1

    if i + 6 >= n:
        break

    if not all(is_number(lines[j]) for j in range(i, i + 3)):
        continue

    gross = float(lines[i]); i += 1
    discount = float(lines[i]); i += 1
    net = float(lines[i]); i += 1

    cgst = float(lines[i + 1]) if is_number(lines[i + 1]) else 0
    sgst = float(lines[i + 3]) if is_number(lines[i + 3]) else 0
    i += 4

    if not is_number(lines[i]):
        continue

    total = float(lines[i])
    i += 1

    items.append({
        "Description": description,
        "UnitPrice": gross,
        "Discount": discount,
        "Qty": 1,
        "NetAmount": net,
        "TaxRate": 5,
        "TaxType": "GST",
        "TaxAmount": round(cgst + sgst, 2),
        "TotalAmount": total
    })

df_items = pd.DataFrame(items)

# ================= WRITE TO EXCEL =================
wb = load_workbook(TEMPLATE_PATH)

# ---- Sheet 1 (Header) ----
ws1 = wb.active
for row in range(2, ws1.max_row + 1):
    field = ws1.cell(row=row, column=1).value
    if field in data:
        ws1.cell(row=row, column=2).value = data[field]

# ---- Table_1 (Line Items) ----
ws = wb["Table_1"]
start_row = 2

for idx, row in df_items.iterrows():
    r = start_row + idx
    ws.cell(r, 1).value = idx + 1
    ws.cell(r, 2).value = row["Description"]
    ws.cell(r, 3).value = row["UnitPrice"]
    ws.cell(r, 4).value = row["Discount"]
    ws.cell(r, 5).value = row["Qty"]
    ws.cell(r, 6).value = row["NetAmount"]
    ws.cell(r, 7).value = row["TaxRate"]
    ws.cell(r, 8).value = row["TaxType"]
    ws.cell(r, 9).value = row["TaxAmount"]
    ws.cell(r, 10).value = row["TotalAmount"]

wb.save(TEMPLATE_PATH)

print("Zomato invoice processed successfully and Table_1 filled")
