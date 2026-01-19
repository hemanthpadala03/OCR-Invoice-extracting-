import easyocr
import re
import pandas as pd
import cv2
from openpyxl import load_workbook
import os

# =====================================================
# PATHS
# =====================================================
IMAGE_PATH = r"C:\Drive_d\Python\F-AI\T4\Input\images\ForwardInvoice_ORD66093937402_page_1.png"
TEMPLATE_PATH = r"C:\Drive_d\Python\F-AI\T4\Output Template.xlsx"
OUTPUT_PATH = r"C:\Drive_d\Python\F-AI\T4\Outputs\Blinkit\Blinkit_Output.xlsx"

TABLE_SHEET  = "Table_1"
HEADER_SHEET = "Invoice_Header"

os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

# =====================================================
# OCR INIT
# =====================================================
reader = easyocr.Reader(["en"], gpu=False)

# =====================================================
# OCR – FULL IMAGE (FOR TABLE)
# =====================================================
full_results = reader.readtext(IMAGE_PATH, detail=0)
full_lines = [l.strip() for l in full_results if l.strip()]
full_text = "\n".join(full_lines)

# =====================================================
# OCR – SPLIT IMAGE (FOR HEADER)
# =====================================================
img = cv2.imread(IMAGE_PATH)
h, w, _ = img.shape
split_x = int(w * 0.65)

left_img  = img[:, :split_x]
right_img = img[:, split_x:]

left_text  = "\n".join([t for _, t, _ in reader.readtext(left_img)])
right_text = "\n".join([t for _, t, _ in reader.readtext(right_img)])
header_text = left_text + "\n" + right_text

# =====================================================
# HELPERS
# =====================================================
NUM_RE = r"\d+\.\d+"

def grab(pattern, text):
    m = re.search(pattern, text, re.I | re.S)
    return m.group(1).strip() if m and m.lastindex else (m.group(0).strip() if m else "")

def clean_price(val):
    if val.startswith("8") and len(val) > 6:
        val = val[1:]
    return float(val)

# =====================================================
# ===================== TABLE_1 =======================
# =====================================================
numbers = re.findall(NUM_RE, full_text)
numbers = [clean_price(n) for n in numbers]

"""
Expected Blinkit numeric order (stable):
[MRP, Discount, Taxable, CGST%, CGST_AMT, SGST%, SGST_AMT, Cess, AddCess, Total]
Example:
[135.00, 33.00, 97.14, 2.50, 2.43, 2.50, 2.43, 0.00, 0.00, 102.00]
"""

if len(numbers) < 8:
    raise RuntimeError("❌ Blinkit table values not detected (numbers too few)")

mrp         = numbers[0]
discount    = numbers[1]
taxable_val = numbers[2]

# tax values (robust)
tax_amounts = [n for n in numbers if 0 < n < 10]
total_tax = round(sum(tax_amounts[:2]), 2)

total_amt = numbers[-1]
net_amount = taxable_val

# description (text between item name & numbers)
desc_lines = []
for l in full_lines:
    if re.search(r"(Dairy Day|Ice Cream|Handling charge)", l, re.I):
        desc_lines.append(l)
    elif desc_lines and re.search(r"\d+\.\d+", l):
        break
    elif desc_lines:
        desc_lines.append(l)

description = " ".join(desc_lines).strip()

rows = [[
    1,
    description,
    mrp,
    discount,
    1,
    net_amount,
    "5%",
    "CGST+SGST",
    total_tax,
    total_amt
]]

df = pd.DataFrame(rows, columns=[
    "Sl.No",
    "Description",
    "UnitPrice",
    "Discount",
    "Qty",
    "NetAmount",
    "TaxRate",
    "TaxType",
    "TaxAmount",
    "TotalAmount"
])

# =====================================================
# WRITE TABLE_1
# =====================================================
wb = load_workbook(TEMPLATE_PATH)
ws_table = wb[TABLE_SHEET]

start_row = 2
for r, row in enumerate(df.itertuples(index=False), start_row):
    for c, val in enumerate(row, 1):
        ws_table.cell(row=r, column=c).value = val

total_row = start_row + len(df)
ws_table.cell(row=total_row, column=2).value = "TOTAL:"
ws_table.cell(row=total_row, column=9).value = total_tax
ws_table.cell(row=total_row, column=10).value = total_amt

# =====================================================
# ================= INVOICE_HEADER ====================
# =====================================================
header_data = {
    "billing_address": grab(
        r"Invoice To\s*(.*?)\s*Sr\.?\s*no",
        full_text
    ),

    "shipping_address": grab(
        r"Invoice To\s*(.*?)\s*Sr\.?\s*no",
        full_text
    ),



    "invoice_type": "Tax Invoice",

    "order_number": grab(
        r"Order Id\s*(\d+)",
        full_text
    ),

    "invoice_number": grab(
        r"Invoice Number\s*([A-Z0-9]+)",
        full_text
    ),

    "order_date": grab(
        r"Invoice\s*(\d{2}-[A-Za-z]{3}-\d{4})",
        full_text
    ),

    "invoice_details": grab(
        r"Invoice Number\s*[A-Z0-9]+",
        full_text
    ),

    "invoice_date": grab(
        r"Invoice\s*(\d{2}-[A-Za-z]{3}-\d{4})",
        full_text
    ),

    "seller_info": grab(
        r"Seller\s*(Zomato Hyperpure Private Limited.*?)GSTIN",
        full_text
    ),

    "seller_pan": grab(
        r"PAN\s*([A-Z0-9]+)",
        full_text
    ),

    "seller_gst": grab(
        r"GSTIN\s*(\d{2}[A-Z0-9]+)",
        full_text
    ),

    "fssai_license": grab(
        r"FSSAI License Number\s*(\d+)",
        full_text
    ),

    "billing_state_code": grab(
        r"State\s*(Tamil Nadu)",
        full_text
    ),

    "shipping_state_code": grab(
        r"State\s*(Tamil Nadu)",
        full_text
    ),

    "place_of_supply": grab(
        r"Place of\s*Supply\s*(Tamil Nadu)",
        full_text
    ),

    "place_of_delivery": grab(
        r"Place of\s*Supply\s*(Tamil Nadu)",
        full_text
    ),

    "reverse_charge": grab(
        r"reverse charge\s*\n*(Yes|No)",
        full_text
    ),

    "amount_in_words": grab(
        r"Amount in\s*One\s*(.*?)\s*Words",
        full_text
    ),

    "seller_name": "Zomato Hyperpure Private Limited",

    "seller_address": grab(
        r"Zomato Hyperpure Private Limited\s*(.*?)GSTIN",
        full_text
    ),

    "total_tax": total_tax,

    "total_amount": total_amt
}


ws_header = wb[HEADER_SHEET]
row = 2
for k, v in header_data.items():
    ws_header.cell(row=row, column=1).value = k
    ws_header.cell(row=row, column=2).value = v
    row += 1

wb.save(OUTPUT_PATH)

print("✅ BLINKIT INVOICE PARSED SUCCESSFULLY")
print("📁 Output:", OUTPUT_PATH)
