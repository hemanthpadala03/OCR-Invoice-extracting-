import easyocr
import re
import pandas as pd
import cv2     
from openpyxl import load_workbook
IMAGE_PATH = r"T4/Input/images/225790297864463_page_1.png"   # change this
TEMPLATE_PATH = r"C:\Drive_d\Python\F-AI\T4\Output Template.xlsx"
OUTPUT_PATH = r"C:\Drive_d\Python\F-AI\T4\Output.xlsx"
 
img = cv2.imread(IMAGE_PATH)
h, w, _ = img.shape
left_col  = img[:, :w//2]
right_col = img[:, w//2:]
reader = easyocr.Reader(['en'], gpu=False)

# print("Full Image")
results = reader.readtext(IMAGE_PATH)


lines = [t.strip() for _, t, _ in results if t.strip()]

def is_num(x):
    return re.fullmatch(r"\d+(\.\d+)?", x) is not None

def to_f(x):
    try:
        return float(x)
    except:
        return 0.0

products = []
i = 0
n = len(lines)

while i < n and lines[i] != "(Rs.)":
    i += 1
i += 1

while i < n:
    if "invoice value" in lines[i].lower():
        break

    name_parts = []

    while i < n and not is_num(lines[i]) and lines[i] != "NOS":
        name_parts.append(lines[i])
        i += 1

    if i < n and is_num(lines[i]):
        qty = int(float(lines[i]))
        i += 1
    else:
        qty = 1

    if i < n and lines[i] == "NOS":
        i += 1

    hsn = lines[i] if i < n else "0"
    i += 1

    taxable = to_f(lines[i]) if i < n else 0
    i += 1

    discount = to_f(lines[i]) if i < n else 0
    i += 1

    net_taxable = to_f(lines[i]) if i < n else 0
    i += 1

    cgst_pct = cgst_amt = sgst_pct = sgst_amt = 0

    if i < n and is_num(lines[i]):
        cgst_pct = to_f(lines[i])
        i += 1

    if i < n and is_num(lines[i]):
        cgst_amt = to_f(lines[i])
        i += 1

    if i < n and is_num(lines[i]):
        sgst_pct = to_f(lines[i])
        i += 1

    if i < n and is_num(lines[i]):
        sgst_amt = to_f(lines[i])
        i += 1

    line_total = to_f(lines[i]) if i < n else 0
    i += 1

    while i < n and not is_num(lines[i]) and lines[i] != "NOS":
        if "invoice value" in lines[i].lower():
            break
        name_parts.append(lines[i])
        i += 1

    products.append({
        "Description": " ".join(name_parts),
        "Quantity": qty,
        "UQC": "NOS",
        "HSN": hsn,
        "Taxable Value": taxable,
        "Discount": discount,
        "Net Taxable Value": net_taxable,
        "CGST %": cgst_pct,
        "CGST Amount": cgst_amt,
        "SGST %": sgst_pct,
        "SGST Amount": sgst_amt,
        "Total Amount": line_total
    })

df = pd.DataFrame(products)
# print(df)

# df.to_excel("products.xlsx", index=False)

with pd.ExcelWriter(
    r"C:\Drive_d\Python\F-AI\T4\Output Template.xlsx",
    engine="openpyxl",
    mode="a",
    if_sheet_exists="replace"
) as writer:
    df.to_excel(writer, sheet_name="Sheet2", index=False)







# for bbox, text, conf in results:
#     print(f"{text}")
    

# print("left col")
# Left_results = reader.readtext(left_col)
# for bbox, text, conf in Left_results:
#     print(f"{text}")

# print("\nright col")
# Right_results = reader.readtext(right_col)      
# for bbox, text, conf in Right_results:
#     print(f"{text}")

left_lines = [t.strip() for _, t, _ in reader.readtext(left_col) if t.strip()]
right_lines = [t.strip() for _, t, _ in reader.readtext(right_col) if t.strip()]
full_text = "\n".join(left_lines + right_lines)

def clean(s):
    return re.sub(r"\s+", " ", s).strip()

data = {}

# ---------------- BILLING / SHIPPING ADDRESS ----------------
addr_text = "\n".join(left_lines)
addr_match = re.search(
    r"(202,\s*Kasa.*?India)",
    addr_text,
    re.I | re.S
)
billing_addr = clean(addr_match.group(1)) if addr_match else ""

data["billing_address"] = billing_addr
data["shipping_address"] = billing_addr

# ---------------- INVOICE META ----------------
data["invoice_type"] = "TAX IN"

data["order_number"] = clean(
    re.search(r"Order\s*ID[:\-]?\s*(\d+)", full_text, re.I).group(1)
)

data["invoice_number"] = clean(
    re.search(r"Invoice\s*No[:\-]?\s*([A-Z0-9]+)", full_text, re.I).group(1)
)

date_match = re.search(r"Date\s*of\s*Invoice[:\-]?\s*([\d\-\/]+)", full_text, re.I)
data["order_date"] = date_match.group(1)
data["invoice_date"] = date_match.group(1)

data["invoice_details"] = "GST Invoice"

# ---------------- SELLER NAME (HARD STOP) ----------------
seller_name = ""
for i, line in enumerate(right_lines):
    if "seller name" in line.lower():
        seller_name = right_lines[i + 1]
        break
data["seller_name"] = clean(seller_name)

# ---------------- SELLER ADDRESS ----------------
seller_addr = []
addr_start = False
for line in right_lines:
    if "address" in line.lower():
        addr_start = True
        continue
    if addr_start:
        if any(k in line.lower() for k in ["state", "gstin", "fssai"]):
            break
        seller_addr.append(line)

data["seller_address"] = clean(" ".join(seller_addr))

# ---------------- GST / PAN / FSSAI ----------------
gst = re.search(r"\b[0-9A-Z]{15}\b", full_text)
data["seller_gst"] = gst.group(0) if gst else ""
data["seller_pan"] = data["seller_gst"][2:12] if data["seller_gst"] else ""

fssai = re.search(r"\b\d{14}\b", full_text)
data["fssai_license"] = fssai.group(0) if fssai else ""

# ---------------- PLACE / STATE ----------------
data["place_of_supply"] = "Telangana"
data["place_of_delivery"] = "Telangana"

pin = re.search(r"\b502\d{3}\b", full_text)
data["billing_state_code"] = pin.group(0) if pin else ""
data["shipping_state_code"] = data["billing_state_code"]

data["reverse_charge"] = "No"

# ---------------- AMOUNT IN WORDS ----------------
words_match = re.search(
    r"Amount\s*in\s*words[:\-]?\s*(.*?)Only",
    full_text,
    re.I | re.S
)
data["amount_in_words"] = clean(words_match.group(1)) if words_match else ""

# ---------------- TOTAL AMOUNT & TAX ----------------
total_amt = re.search(r"Invoice\s*Value\s*(\d+)", full_text)
data["total_amount"] = total_amt.group(1) if total_amt else ""

try:
    data["total_tax"] = round(float(data["total_amount"]) * 0.05, 2)
except:
    data["total_tax"] = ""

data["seller_info"] = f"{data['seller_name']} | GST: {data['seller_gst']}"

# ---------------- WRITE BACK TO TEMPLATE ----------------
wb = load_workbook(TEMPLATE_PATH)
ws = wb.active

for row in range(2, ws.max_row + 1):
    field = ws.cell(row=row, column=1).value
    if field in data:
        ws.cell(row=row, column=2).value = data[field]

wb.save(TEMPLATE_PATH)

print("Output Template updated with cleaned values")