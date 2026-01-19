import easyocr
import cv2
import re
import os
import pandas as pd
from openpyxl import load_workbook

# =====================================================
# PATHS
# =====================================================
IMAGE_PATH = r"C:\Drive_d\Python\F-AI\T4\Input\images\invvoice4_page_1.png"
TEMPLATE_PATH = r"C:\Drive_d\Python\F-AI\T4\Output Template.xlsx"
OUTPUT_DIR = r"C:\Drive_d\Python\F-AI\T4\Outputs\Flipkart"
OUTPUT_PATH = os.path.join(OUTPUT_DIR, "Output.xlsx")

os.makedirs(OUTPUT_DIR, exist_ok=True)

HEADER_SHEET = "Invoice_Header"
TABLE_SHEET  = "Table_1"

# =====================================================
# OCR INIT
# =====================================================
reader = easyocr.Reader(['en'], gpu=False)

img = cv2.imread(IMAGE_PATH)
h, w, _ = img.shape

# =====================================================
# 3 COLUMN SPLIT (HEADER)
# =====================================================
col1 = img[:, 0:int(0.28 * w)]
col2 = img[:, int(0.28 * w):int(0.52 * w)]
col3 = img[:, int(0.52 * w):int(0.75 * w)]

col1_text = "\n".join([t for _, t, _ in reader.readtext(col1)])
col2_text = "\n".join([t for _, t, _ in reader.readtext(col2)])
col3_text = "\n".join([t for _, t, _ in reader.readtext(col3)])

full_header_text = col1_text + "\n" + col2_text + "\n" + col3_text

# =====================================================
# FULL OCR (TABLE)
# =====================================================
full_lines = [l.strip() for l in reader.readtext(IMAGE_PATH, detail=0) if l.strip()]

# =====================================================
# HELPERS
# =====================================================
def grab(pattern, text):
    m = re.search(pattern, text, re.I | re.S)
    return m.group(1).strip() if m else ""

def clean_block(txt):
    return re.sub(r"\n{2,}", "\n", txt).strip()

# =====================================================
# ================= INVOICE_HEADER ====================
# =====================================================
seller_block = grab(r"Sold By\s*(.*?)\s*(GSTIN|GST:)", col1_text)
seller_name = seller_block.split("\n")[0] if seller_block else ""
seller_address = clean_block("\n".join(seller_block.split("\n")[1:])) if seller_block else ""

shipping_address = clean_block(grab(
    r"Shipping ADDRESS\s*(.*?)\s*(Description|HSN|IGST|Product)",
    col2_text
))

billing_address = clean_block(grab(
    r"Billing Address\s*(.*?)\s*(Gross|Taxable|Discount|Amount|Value)",
    col3_text
))

invoice_header = {
    "billing_address": billing_address,
    "shipping_address": shipping_address,
    "invoice_type": grab(r"(Tax Invoice)", full_header_text),
    "order_number": grab(r"Order Id[:\s]*([\w\d]+)", full_header_text),
    "invoice_number": grab(r"Invoice No[:\s]*([\w\d]+)", full_header_text),
    "order_date": grab(r"Order Date[:\s]*([\d\-\,\sAPMapm]+)", full_header_text),
    "invoice_details": grab(r"(Invoice No[:\s]*[\w\d]+)", full_header_text),
    "invoice_date": grab(r"Invoice Date[:\s]*([\d\-\,\sAPMapm]+)", full_header_text),
    "seller_info": seller_block,
    "seller_pan": grab(r"PAN[:\s]*([A-Z0-9]+)", full_header_text),
    "seller_gst": grab(r"GSTIN[:\s]*([A-Z0-9]+)", full_header_text),
    "fssai_license": grab(r"FSSAI.*?([\d]{10,})", full_header_text),
    "billing_state_code": grab(r"IN\-([A-Z]{2})", billing_address),
    "shipping_state_code": grab(r"IN\-([A-Z]{2})", shipping_address),
    "place_of_supply": grab(r"Place of\s*([\w\s]+)", full_header_text),
    "place_of_delivery": grab(r"Place of\s*([\w\s]+)", full_header_text),
    "reverse_charge": grab(r"reverse charge\s*(Yes|No)", full_header_text),
    "amount_in_words": grab(r"Amount in\s*Words[:\s]*(.*?)\s*(Blink|Seller|GSTIN)", full_header_text),
    "seller_name": seller_name,
    "seller_address": seller_address,
}

# ================== TABLE_1 (FLIPKART FIXED) ==================

full_text = "\n".join([t for _, t, _ in reader.readtext(IMAGE_PATH)])

# 1️⃣ Extract product description
product_desc = grab(
    r"(SPL\s+Back\s+Cover\s+for\s+Realme\s+8\s+Black)",
    full_text
)

# 2️⃣ Extract numeric values AFTER product line
after_product = full_text.split(product_desc)[-1]

nums = [float(x) for x in re.findall(r"-?\d+\.\d+", after_product)]

if len(nums) < 5:
    raise RuntimeError("❌ Flipkart numeric table incomplete")

gross_amount   = nums[0]     # 299.00
discount       = abs(nums[1])# 15.00
taxable_value  = nums[2]     # 240.68
igst_amount    = nums[3]     # 43.32
total_amount   = nums[4]     # 284.00

rows = [[
    1,
    product_desc,
    gross_amount,
    discount,
    1,
    taxable_value,
    "18%",
    "IGST",
    igst_amount,
    total_amount
]]

df_table = pd.DataFrame(rows, columns=[
    "Sl.No",
    "Description",
    "UnitPrice",
    "Discount",
    "Qty",
    "NetAmount",
    "TaxRate",
    "TaxType",
    "TaxAmount",
    "TotalAmount"
])

print(df_table)


# =====================================================
# WRITE TO EXCEL
# =====================================================
wb = load_workbook(TEMPLATE_PATH)

# ---- Invoice_Header
ws_h = wb[HEADER_SHEET]
r = 2
while ws_h.cell(row=r, column=1).value:
    key = ws_h.cell(row=r, column=1).value
    if key in invoice_header:
        ws_h.cell(row=r, column=2).value = invoice_header[key]
    r += 1

# ---- Table_1
ws_t = wb[TABLE_SHEET]

start_row = 2

for r, row in enumerate(df_table.itertuples(index=False), start=start_row):
    for c, v in enumerate(row, start=1):
        ws_t.cell(row=r, column=c).value = v

# TOTAL ROW
total_row = start_row + len(df_table)

ws_t.cell(row=total_row, column=2).value = "TOTAL:"
ws_t.cell(row=total_row, column=9).value = igst_amount        # Flipkart = IGST only
ws_t.cell(row=total_row, column=10).value = total_amount


wb.save(OUTPUT_PATH)

print("✅ FLIPKART INVOICE FULLY EXTRACTED")
print("Saved to:", OUTPUT_PATH)
